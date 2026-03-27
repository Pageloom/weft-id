"""Tests for bulk user template export job handler."""

from io import BytesIO
from unittest.mock import MagicMock, patch

import database
import pytest
from openpyxl import load_workbook


class TestHandleExportUsersTemplate:
    """Tests for the export_users_template job handler."""

    def test_generates_xlsx_with_users(self, test_tenant, test_admin_user):
        from jobs.export_users_template import handle_export_users_template

        bg_task = database.bg_tasks.create_task(
            tenant_id=test_tenant["id"],
            job_type="export_users_template",
            created_by=test_admin_user["id"],
        )

        task = {
            "id": bg_task["id"],
            "tenant_id": test_tenant["id"],
            "created_by": test_admin_user["id"],
            "job_type": "export_users_template",
            "payload": None,
        }

        saved_data = {}

        def capture_save(key, data, content_type):
            saved_data["key"] = key
            saved_data["data"] = data.read()
            saved_data["content_type"] = content_type
            return f"local/{key}"

        with patch("jobs.export_users_template.storage.get_backend") as mock_get_backend:
            mock_backend = MagicMock()
            mock_backend.save.side_effect = capture_save
            mock_get_backend.return_value = mock_backend

            result = handle_export_users_template(task)

        assert result["records_processed"] >= 1
        assert result["filename"].endswith(".xlsx")
        assert test_tenant["subdomain"] in result["filename"]
        assert result["file_id"] is not None

        # Verify XLSX content
        wb = load_workbook(filename=BytesIO(saved_data["data"]))
        ws = wb.active
        assert ws.title == "Users"

        # Verify header
        header = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert header == [
            "user_id",
            "email",
            "domain",
            "first_name",
            "last_name",
            "new_secondary_email",
            "new_first_name",
            "new_last_name",
        ]

        # Verify at least one data row
        assert ws.cell(row=2, column=1).value is not None

        # Verify content type
        assert saved_data["content_type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        wb.close()

    def test_empty_tenant_generates_header_only(self, test_tenant, test_admin_user):
        """A tenant with no active users should still produce a valid XLSX."""

        from jobs.export_users_template import handle_export_users_template

        bg_task = database.bg_tasks.create_task(
            tenant_id=test_tenant["id"],
            job_type="export_users_template",
            created_by=test_admin_user["id"],
        )

        task = {
            "id": bg_task["id"],
            "tenant_id": test_tenant["id"],
            "created_by": test_admin_user["id"],
            "job_type": "export_users_template",
            "payload": None,
        }

        saved_data = {}

        def capture_save(key, data, content_type):
            saved_data["data"] = data.read()
            return f"local/{key}"

        with (
            patch("jobs.export_users_template.storage.get_backend") as mock_get_backend,
            patch(
                "jobs.export_users_template.database.users.list_all_users_for_export"
            ) as mock_list,
        ):
            mock_list.return_value = []
            mock_backend = MagicMock()
            mock_backend.save.side_effect = capture_save
            mock_get_backend.return_value = mock_backend

            result = handle_export_users_template(task)

        assert result["records_processed"] == 0

        # Should still have valid XLSX with header
        wb = load_workbook(filename=BytesIO(saved_data["data"]))
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "user_id"
        assert ws.cell(row=2, column=1).value is None
        wb.close()

    def test_nonexistent_tenant_raises(self, test_admin_user):
        from uuid import uuid4

        from jobs.export_users_template import handle_export_users_template

        task = {
            "id": uuid4(),
            "tenant_id": uuid4(),
            "created_by": test_admin_user["id"],
            "job_type": "export_users_template",
            "payload": None,
        }

        with pytest.raises(ValueError, match="does not exist"):
            handle_export_users_template(task)

    def test_export_file_record_created(self, test_tenant, test_admin_user):
        from jobs.export_users_template import handle_export_users_template

        bg_task = database.bg_tasks.create_task(
            tenant_id=test_tenant["id"],
            job_type="export_users_template",
            created_by=test_admin_user["id"],
        )

        task = {
            "id": bg_task["id"],
            "tenant_id": test_tenant["id"],
            "created_by": test_admin_user["id"],
            "job_type": "export_users_template",
            "payload": None,
        }

        with patch("jobs.export_users_template.storage.get_backend") as mock_get_backend:
            mock_backend = MagicMock()
            mock_backend.save.return_value = "local/test"
            mock_get_backend.return_value = mock_backend

            result = handle_export_users_template(task)

        # Verify export file was recorded in database
        export = database.export_files.get_export_file(str(test_tenant["id"]), result["file_id"])
        assert export is not None
        assert export["filename"] == result["filename"]
        assert export["content_type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
