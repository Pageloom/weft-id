"""User audit export job handler.

Exports all users, group memberships, and app access as a
password-encrypted multi-sheet XLSX workbook.
"""

import logging
from collections import defaultdict
from datetime import UTC, datetime, timedelta
from typing import Any
from uuid import uuid4

import database
import settings
from jobs.registry import register_handler
from openpyxl import Workbook
from openpyxl.styles import Font
from utils import storage
from utils.xlsx_encryption import encrypt_workbook

logger = logging.getLogger(__name__)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _user_status(user: dict) -> str:
    if user["is_anonymized"]:
        return "anonymized"
    if user["is_inactivated"]:
        return "inactive"
    return "active"


def _auth_method(user: dict) -> str:
    if user.get("saml_idp_name"):
        return str(user["saml_idp_name"])
    if user.get("has_password"):
        return "password"
    return "none"


def _fmt_dt(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S UTC")
    return str(val)


def _build_users_sheet(
    wb: Workbook,
    users: list[dict],
    secondary_map: dict[str, list[str]],
    creation_map: dict[str, str],
    login_ip_map: dict[str, str],
    app_count_map: dict[str, int],
) -> None:
    """Build Sheet 1: Users (one row per user)."""
    ws = wb.active
    ws.title = "Users"

    headers = [
        "User ID",
        "First Name",
        "Last Name",
        "Primary Email",
        "Domain",
        "Secondary Emails",
        "Role",
        "Status",
        "Created",
        "Creation Method",
        "Auth Method",
        "Last Login",
        "Last Login IP",
        "Last Activity",
        "Password Changed",
        "MFA Enabled",
        "App Count",
    ]
    ws.append(headers)
    header_font = Font(bold=True, size=14)
    for cell in ws[1]:
        cell.font = header_font

    for user in users:
        uid = str(user["id"])
        email = user.get("primary_email") or ""
        domain = email.split("@")[1] if "@" in email else ""
        secondaries = ", ".join(secondary_map.get(uid, []))

        ws.append(
            [
                uid,
                user.get("first_name") or "",
                user.get("last_name") or "",
                email,
                domain,
                secondaries,
                user["role"],
                _user_status(user),
                _fmt_dt(user.get("created_at")),
                creation_map.get(uid, "cli"),
                _auth_method(user),
                _fmt_dt(user.get("last_login")),
                login_ip_map.get(uid, ""),
                _fmt_dt(user.get("last_activity_at")),
                _fmt_dt(user.get("password_changed_at")),
                "Yes" if user.get("mfa_enabled") else "No",
                app_count_map.get(uid, 0),
            ]
        )

    ws.auto_filter.ref = ws.dimensions


def _build_memberships_sheet(
    wb: Workbook,
    memberships: list[dict],
) -> None:
    """Build Sheet 2: Group Memberships (one row per user-group pair)."""
    ws = wb.create_sheet("Group Memberships")

    headers = [
        "User ID",
        "Email",
        "Group Name",
        "Group Type",
        "Member Since",
    ]
    ws.append(headers)
    header_font = Font(bold=True, size=14)
    for cell in ws[1]:
        cell.font = header_font

    for m in memberships:
        ws.append(
            [
                str(m["user_id"]),
                m.get("email") or "",
                m["group_name"],
                m["group_type"],
                _fmt_dt(m.get("membership_since")),
            ]
        )

    ws.auto_filter.ref = ws.dimensions


def _build_access_sheet(
    wb: Workbook,
    access_rows: list[dict],
    assertion_map: dict[tuple[str, str], datetime],
) -> None:
    """Build Sheet 3: App Access (one row per user-SP pair)."""
    ws = wb.create_sheet("App Access")

    headers = [
        "User ID",
        "Email",
        "App Name",
        "Last Auth",
        "Access Via",
    ]
    ws.append(headers)
    header_font = Font(bold=True, size=14)
    for cell in ws[1]:
        cell.font = header_font

    # Aggregate access_rows: group by (user_id, sp_id) to collect group names
    aggregated: dict[tuple[str, str], dict] = {}
    for row in access_rows:
        uid = str(row["user_id"])
        sp_id = str(row["sp_id"])
        key = (uid, sp_id)

        if key not in aggregated:
            aggregated[key] = {
                "user_id": uid,
                "email": row.get("email") or "",
                "app_name": row["app_name"],
                "groups": set(),
                "available_to_all": False,
            }

        if row["available_to_all"]:
            aggregated[key]["available_to_all"] = True
        elif row.get("granting_group_name"):
            aggregated[key]["groups"].add(row["granting_group_name"])

    # Sort by email then app name for consistent output
    sorted_rows = sorted(aggregated.values(), key=lambda r: (r["email"], r["app_name"]))

    for entry in sorted_rows:
        uid = entry["user_id"]
        sp_id_for_lookup = None
        # Find sp_id from the original data for assertion lookup
        for row in access_rows:
            if str(row["user_id"]) == uid and row["app_name"] == entry["app_name"]:
                sp_id_for_lookup = str(row["sp_id"])
                break

        last_auth = ""
        if sp_id_for_lookup:
            auth_dt = assertion_map.get((uid, sp_id_for_lookup))
            if auth_dt:
                last_auth = _fmt_dt(auth_dt)

        if entry["available_to_all"]:
            access_via = "All users"
        else:
            access_via = ", ".join(sorted(entry["groups"]))

        ws.append(
            [
                uid,
                entry["email"],
                entry["app_name"],
                last_auth,
                access_via,
            ]
        )

    ws.auto_filter.ref = ws.dimensions


@register_handler("export_users")
def handle_export_users(task: dict) -> dict[str, Any]:
    """Export user audit data as a password-encrypted multi-sheet XLSX.

    Produces three sheets: Users, Group Memberships, and App Access.

    Args:
        task: Dict with id, tenant_id, job_type, created_by.

    Returns:
        Dict with output, file_id, password, records_processed,
        filename, file_size.
    """
    tenant_id = str(task["tenant_id"])
    created_by = str(task["created_by"])
    task_id = str(task["id"])

    logger.info("Starting user audit XLSX export for tenant %s", tenant_id)

    tenant = database.tenants.get_tenant_by_id(tenant_id)
    if not tenant:
        raise ValueError(f"Tenant {tenant_id} does not exist")

    # Fetch all data via batch queries
    users = database.users.get_all_users_for_audit(tenant_id)
    logger.info("Fetched %d users", len(users))

    secondary_emails_raw = database.users.get_all_secondary_emails(tenant_id)
    secondary_map: dict[str, list[str]] = defaultdict(list)
    for row in secondary_emails_raw:
        secondary_map[str(row["user_id"])].append(row["email"])

    creation_raw = database.users.get_creation_methods(tenant_id)
    creation_map: dict[str, str] = {}
    for row in creation_raw:
        event_type = row["event_type"]
        if event_type == "user_created_jit":
            creation_map[str(row["user_id"])] = "jit"
        else:
            creation_map[str(row["user_id"])] = "invited"

    login_ip_raw = database.users.get_last_login_ips(tenant_id)
    login_ip_map: dict[str, str] = {}
    for row in login_ip_raw:
        metadata = row.get("metadata") or {}
        ip = metadata.get("remote_address", "")
        if ip:
            login_ip_map[str(row["user_id"])] = ip

    app_count_raw = database.users.get_app_counts(tenant_id)
    app_count_map: dict[str, int] = {}
    for row in app_count_raw:
        app_count_map[str(row["user_id"])] = row["app_count"]

    # Add available_to_all SP count to every user
    ata_count = database.users.get_available_to_all_sp_count(tenant_id)
    if ata_count > 0:
        user_ids = {str(u["id"]) for u in users}
        for uid in user_ids:
            app_count_map[uid] = app_count_map.get(uid, 0) + ata_count

    memberships = database.users.get_all_group_memberships(tenant_id)
    logger.info("Fetched %d group memberships", len(memberships))

    access_rows = database.users.get_all_user_sp_access(tenant_id)
    logger.info("Fetched %d SP access rows", len(access_rows))

    assertion_raw = database.users.get_last_sso_assertions(tenant_id)
    assertion_map: dict[tuple[str, str], datetime] = {}
    for row in assertion_raw:
        assertion_map[(str(row["user_id"]), str(row["sp_id"]))] = row["last_auth_at"]

    # Build workbook
    wb = Workbook()
    wb._named_styles["Normal"].font = Font(name="Calibri", size=14)

    _build_users_sheet(wb, users, secondary_map, creation_map, login_ip_map, app_count_map)
    _build_memberships_sheet(wb, memberships)
    _build_access_sheet(wb, access_rows, assertion_map)

    # Encrypt
    encrypted = encrypt_workbook(wb)

    # Save
    ts = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    suffix = uuid4().hex[:8]
    filename = f"user-audit_{ts}_{suffix}.xlsx"
    storage_key = f"exports/{tenant_id}/{filename}"

    backend = storage.get_backend()
    backend.save(storage_key, encrypted.data, XLSX_CONTENT_TYPE)
    storage_type = settings.STORAGE_BACKEND.lower()
    if storage_type != "spaces" or not settings.SPACES_BUCKET:
        storage_type = "local"

    logger.info("Saved user audit XLSX export: %s", storage_key)

    expires_at = datetime.now(UTC) + timedelta(hours=settings.EXPORT_FILE_EXPIRY_HOURS)

    export_file = database.export_files.create_export_file(
        tenant_id=tenant_id,
        bg_task_id=task_id,
        filename=filename,
        storage_type=storage_type,
        storage_path=storage_key,
        file_size=encrypted.file_size,
        content_type=XLSX_CONTENT_TYPE,
        expires_at=expires_at,
        created_by=created_by,
    )

    logger.info(
        "Created export file record: %s",
        export_file["id"] if export_file else "None",
    )

    size_kb = encrypted.file_size // 1024
    output_msg = (
        f"Exported {len(users):,} users, "
        f"{len(memberships):,} group memberships, "
        f"{len(set((str(r['user_id']), str(r['sp_id'])) for r in access_rows)):,} "
        f"app access entries to {filename} ({size_kb:,} KB encrypted)"
    )
    return {
        "output": output_msg,
        "file_id": str(export_file["id"]) if export_file else None,
        "records_processed": len(users),
        "filename": filename,
        "file_size": encrypted.file_size,
        "password": encrypted.password,
    }
