"""Event log export job handler.

Exports audit log events as a password-encrypted XLSX file with
human-readable descriptions and resolved artifact/actor names.
"""

import json
import logging
from datetime import UTC, date, datetime, timedelta
from typing import Any
from uuid import uuid4

import database
import settings
from constants.event_types import get_event_description
from jobs.registry import register_handler
from openpyxl import Workbook
from openpyxl.styles import Font
from services.event_log import SYSTEM_ACTOR_ID
from utils import storage
from utils.xlsx_encryption import encrypt_workbook

logger = logging.getLogger(__name__)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MAX_EXPORT_ROWS = 1_000_000


def _resolve_artifact_names(tenant_id: str, events: list[dict]) -> dict[tuple[str, str], str]:
    """Resolve artifact IDs to human-readable names.

    Collects unique (artifact_type, artifact_id) pairs and batch-looks up
    names from the appropriate tables. User artifacts are already resolved
    via the LEFT JOIN in list_events (artifact_first_name/last_name/email),
    so they are handled inline during row writing.
    """
    # Collect unique IDs by type (skip users, handled inline)
    by_type: dict[str, set[str]] = {}
    for e in events:
        atype = e["artifact_type"]
        if atype == "user":
            continue
        aid = str(e["artifact_id"])
        by_type.setdefault(atype, set()).add(aid)

    names: dict[tuple[str, str], str] = {}

    for group_id in by_type.get("group", []):
        g = database.groups.get_group_by_id(tenant_id, group_id)
        if g:
            names[("group", group_id)] = g["name"]

    for sp_id in by_type.get("service_provider", []):
        sp = database.service_providers.get_service_provider(tenant_id, sp_id)
        if sp:
            names[("service_provider", sp_id)] = sp["name"]

    for idp_id in by_type.get("saml_identity_provider", []):
        idp = database.saml.get_identity_provider(tenant_id, idp_id)
        if idp:
            names[("saml_identity_provider", idp_id)] = idp["name"]

    for client_id in by_type.get("oauth2_client", []):
        client = database.oauth2.get_client_by_id(tenant_id, client_id)
        if client:
            names[("oauth2_client", client_id)] = client["name"]

    return names


def _resolve_actor_emails(tenant_id: str, events: list[dict]) -> dict[str, str]:
    """Resolve actor user IDs to primary email addresses."""
    unique_ids = {str(e["actor_user_id"]) for e in events}
    emails: dict[str, str] = {}

    for user_id in unique_ids:
        if user_id == SYSTEM_ACTOR_ID:
            emails[user_id] = "system"
            continue
        primary = database.user_emails.get_primary_email(tenant_id, user_id)
        if primary:
            emails[user_id] = primary["email"]

    return emails


def _get_artifact_name(event: dict, artifact_names: dict[tuple[str, str], str]) -> str:
    """Get artifact display name for a single event row."""
    atype = event["artifact_type"]
    if atype == "user":
        first = event.get("artifact_first_name") or ""
        last = event.get("artifact_last_name") or ""
        email = event.get("artifact_email") or ""
        name = f"{first} {last}".strip()
        if name and email:
            return f"{name} ({email})"
        return name or email or ""
    return artifact_names.get((atype, str(event["artifact_id"])), "")


def _get_actor_email(event: dict, actor_emails: dict[str, str]) -> str:
    """Get actor email, with IdP attribution for system actions."""
    actor_id = str(event["actor_user_id"])
    if actor_id == SYSTEM_ACTOR_ID:
        metadata = event.get("metadata") or {}
        if metadata.get("idp_name"):
            return f"IdP: {metadata['idp_name']}"
        return "System"
    return actor_emails.get(actor_id, "")


def _build_filename(start_date: date | None, end_date: date | None) -> str:
    """Generate export filename with date range, timestamp, and uniqueness suffix."""
    ts = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    suffix = uuid4().hex[:8]
    if start_date and end_date:
        return f"audit-log_{start_date.isoformat()}_to_{end_date.isoformat()}_{ts}_{suffix}.xlsx"
    if start_date:
        return f"audit-log_{start_date.isoformat()}_to_present_{ts}_{suffix}.xlsx"
    if end_date:
        return f"audit-log_up-to_{end_date.isoformat()}_{ts}_{suffix}.xlsx"
    return f"audit-log_all_{ts}_{suffix}.xlsx"


@register_handler("export_events")
def handle_export_events(task: dict) -> dict[str, Any]:
    """Export event logs as a password-encrypted XLSX file.

    Args:
        task: Dict with id, tenant_id, job_type, payload, created_by.
              payload may contain start_date and end_date (ISO 8601 strings).

    Returns:
        Dict with output, file_id, password, records_processed, filename, file_size.
    """
    tenant_id = str(task["tenant_id"])
    created_by = str(task["created_by"])
    task_id = str(task["id"])

    logger.info("Starting event log XLSX export for tenant %s", tenant_id)

    # Validate tenant
    tenant = database.tenants.get_tenant_by_id(tenant_id)
    if not tenant:
        raise ValueError(f"Tenant {tenant_id} does not exist")

    # Parse date range from payload
    payload = task.get("payload") or {}
    start_date = date.fromisoformat(payload["start_date"]) if payload.get("start_date") else None
    end_date = date.fromisoformat(payload["end_date"]) if payload.get("end_date") else None

    # Fetch events in batches
    all_events: list[dict] = []
    offset = 0
    batch_size = 1000

    while True:
        events = database.event_log.list_events(
            tenant_id,
            limit=batch_size,
            offset=offset,
            start_date=start_date,
            end_date=end_date,
        )
        if not events:
            break
        all_events.extend(events)
        offset += batch_size
        logger.info("Fetched %d events so far...", len(all_events))

        if len(all_events) > MAX_EXPORT_ROWS:
            raise ValueError(
                f"Export exceeds {MAX_EXPORT_ROWS:,} rows. Please use a narrower date range."
            )

    logger.info("Total events to export: %d", len(all_events))

    # Resolve names
    artifact_names = _resolve_artifact_names(tenant_id, all_events)
    actor_emails = _resolve_actor_emails(tenant_id, all_events)

    # Build workbook with font size 14 as default
    wb = Workbook()
    default_font = Font(name="Calibri", size=14)
    ws = wb.active
    ws.title = "Audit Log"

    headers = [
        "Timestamp",
        "Event Type",
        "Description",
        "Actor Email",
        "Artifact Type",
        "Artifact ID",
        "Artifact Name",
        "IP Address",
        "User Agent",
        "Device",
        "API Client",
        "Metadata",
    ]
    ws.append(headers)

    header_font = Font(bold=True, size=14)
    for cell in ws[1]:
        cell.font = header_font

    for event in all_events:
        metadata = event.get("metadata") or {}
        api_client = metadata.get("api_client_name", "")
        if api_client and metadata.get("api_client_type"):
            api_client = f"{api_client} ({metadata['api_client_type']})"

        row = [
            event["created_at"].strftime("%Y-%m-%d %H:%M:%S UTC"),
            event["event_type"],
            get_event_description(event["event_type"]) or event["event_type"],
            _get_actor_email(event, actor_emails),
            event["artifact_type"],
            str(event["artifact_id"]),
            _get_artifact_name(event, artifact_names),
            metadata.get("remote_address", ""),
            metadata.get("user_agent", ""),
            metadata.get("device", ""),
            api_client,
            json.dumps(metadata, default=str, ensure_ascii=False) if metadata else "",
        ]
        ws.append(row)

    # Apply font size 14 to all data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = default_font

    # Enable auto-filter on the header row
    ws.auto_filter.ref = ws.dimensions

    # Encrypt
    encrypted = encrypt_workbook(wb)

    # Generate filename and save
    filename = _build_filename(start_date, end_date)
    storage_key = f"exports/{tenant_id}/{filename}"

    backend = storage.get_backend()
    backend.save(storage_key, encrypted.data, XLSX_CONTENT_TYPE)
    storage_type = settings.STORAGE_BACKEND.lower()
    if storage_type != "spaces" or not settings.SPACES_BUCKET:
        storage_type = "local"

    logger.info("Saved XLSX export: %s", storage_key)

    # Calculate expiry and record in database
    expires_at = datetime.now(UTC) + timedelta(hours=settings.EXPORT_FILE_EXPIRY_HOURS)

    export_file = database.export_files.create_export_file(
        tenant_id=tenant_id,
        bg_task_id=task_id,
        filename=filename,
        storage_type=storage_type,
        storage_path=storage_key,
        file_size=encrypted.file_size,
        content_type=XLSX_CONTENT_TYPE,
        expires_at=expires_at,
        created_by=created_by,
    )

    logger.info(
        "Created export file record: %s",
        export_file["id"] if export_file else "None",
    )

    size_kb = encrypted.file_size // 1024
    output_msg = f"Exported {len(all_events):,} events to {filename} ({size_kb:,} KB encrypted)"
    return {
        "output": output_msg,
        "file_id": str(export_file["id"]) if export_file else None,
        "records_processed": len(all_events),
        "filename": filename,
        "file_size": encrypted.file_size,
        "password": encrypted.password,
    }
