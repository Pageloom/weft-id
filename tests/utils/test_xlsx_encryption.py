"""Tests for XLSX encryption utility."""

from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook


class TestEncryptWorkbook:
    """Tests for encrypt_workbook()."""

    def test_returns_encrypted_data(self):
        from utils.xlsx_encryption import encrypt_workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "test"

        result = encrypt_workbook(wb)

        assert result.data is not None
        assert result.password is not None
        assert result.file_size > 0
        assert isinstance(result.data, BytesIO)

    def test_encrypted_file_not_readable_as_plain_xlsx(self):
        from utils.xlsx_encryption import encrypt_workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "secret data"

        result = encrypt_workbook(wb)

        with pytest.raises(Exception):
            load_workbook(filename=result.data)

    def test_correct_password_decrypts(self):
        import msoffcrypto
        from utils.xlsx_encryption import encrypt_workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "secret data"
        ws["B1"] = 42

        result = encrypt_workbook(wb)

        # Decrypt with the returned password
        result.data.seek(0)
        file = msoffcrypto.OfficeFile(result.data)
        file.load_key(password=result.password)
        decrypted = BytesIO()
        file.decrypt(decrypted)
        decrypted.seek(0)

        # Should be readable as XLSX with correct content
        wb2 = load_workbook(filename=decrypted)
        ws2 = wb2.active
        assert ws2["A1"].value == "secret data"
        assert ws2["B1"].value == 42
        wb2.close()

    def test_wrong_password_fails(self):
        import msoffcrypto
        from utils.xlsx_encryption import encrypt_workbook

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "test"

        result = encrypt_workbook(wb)

        result.data.seek(0)
        file = msoffcrypto.OfficeFile(result.data)
        file.load_key(password="wrong-password-here")
        decrypted = BytesIO()
        with pytest.raises(Exception):
            file.decrypt(decrypted)

    def test_password_is_six_word_passphrase(self):
        from utils.xlsx_encryption import encrypt_workbook

        wb = Workbook()
        result = encrypt_workbook(wb)

        words = result.password.split("-")
        assert len(words) == 6
        assert all(w.isalpha() and w.islower() for w in words)

    def test_file_size_matches_buffer(self):
        from utils.xlsx_encryption import encrypt_workbook

        wb = Workbook()
        wb.active["A1"] = "test"

        result = encrypt_workbook(wb)

        assert result.file_size == result.data.getbuffer().nbytes


class TestDecryptXlsxData:
    """Tests for decrypt_xlsx_data()."""

    def test_decrypts_with_correct_password(self):
        from utils.xlsx_encryption import decrypt_xlsx_data, encrypt_workbook

        wb = Workbook()
        wb.active["A1"] = "hello"
        encrypted = encrypt_workbook(wb)

        encrypted_bytes = encrypted.data.read()
        decrypted_bytes = decrypt_xlsx_data(encrypted_bytes, encrypted.password)

        wb2 = load_workbook(filename=BytesIO(decrypted_bytes))
        assert wb2.active["A1"].value == "hello"
        wb2.close()

    def test_wrong_password_raises_value_error(self):
        from utils.xlsx_encryption import decrypt_xlsx_data, encrypt_workbook

        wb = Workbook()
        encrypted = encrypt_workbook(wb)
        encrypted_bytes = encrypted.data.read()

        with pytest.raises(ValueError, match="Failed to decrypt"):
            decrypt_xlsx_data(encrypted_bytes, "wrong-pass")

    def test_plain_xlsx_raises_value_error(self):
        from utils.xlsx_encryption import decrypt_xlsx_data

        wb = Workbook()
        buf = BytesIO()
        wb.save(buf)

        with pytest.raises(ValueError, match="Failed to decrypt"):
            decrypt_xlsx_data(buf.getvalue(), "any-password")
