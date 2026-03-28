"""Bulk user attribute update service layer.

Provides:
- Template download task creation (background job)
- Download retrieval for completed template exports
- Upload task creation (saves file, defers processing to worker)
"""

import logging
from io import BytesIO
from typing import Any
from uuid import uuid4

import database
import settings
from services.activity import track_activity
from services.auth import require_admin
from services.event_log import log_event
from services.exceptions import NotFoundError, ValidationError
from services.types import RequestingUser
from utils import storage

logger = logging.getLogger(__name__)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

EXPECTED_COLUMNS = [
    "user_id",
    "email",
    "domain",
    "first_name",
    "last_name",
    "new_secondary_email",
    "new_first_name",
    "new_last_name",
]


def create_download_task(requesting_user: RequestingUser) -> dict | None:
    """Create a background task to generate the user template spreadsheet.

    Authorization: Requires admin or super_admin role.

    Args:
        requesting_user: The user making the request

    Returns:
        Dict with task id and created_at
    """
    require_admin(requesting_user)
    track_activity(requesting_user["tenant_id"], requesting_user["id"])

    result = database.bg_tasks.create_task(
        tenant_id=requesting_user["tenant_id"],
        job_type="export_users_template",
        created_by=requesting_user["id"],
        payload=None,
    )

    if result:
        log_event(
            tenant_id=requesting_user["tenant_id"],
            actor_user_id=requesting_user["id"],
            artifact_type="bg_task",
            artifact_id=str(result["id"]),
            event_type="bulk_update_task_created",
            metadata={"job_type": "export_users_template"},
        )

    return result


def get_download(
    requesting_user: RequestingUser,
    job_id: str,
) -> dict[str, Any]:
    """Get download info for a completed template export job.

    Authorization: Requires admin. Job must belong to requesting user.

    Args:
        requesting_user: The user making the request
        job_id: The background task ID

    Returns:
        Dict with download info (storage_type, url/path, filename, content_type)

    Raises:
        NotFoundError: If job not found or has no file
        ValidationError: If job is still pending/processing or failed
    """
    require_admin(requesting_user)
    track_activity(requesting_user["tenant_id"], requesting_user["id"])

    tenant_id = requesting_user["tenant_id"]

    task = database.bg_tasks.get_task_for_user(
        tenant_id=tenant_id,
        user_id=requesting_user["id"],
        task_id=job_id,
    )

    if not task:
        raise NotFoundError(
            message="Job not found",
            code="job_not_found",
        )

    if task["status"] in ("pending", "processing"):
        raise ValidationError(
            message="Job is still processing",
            code="job_pending",
            details={"status": task["status"]},
        )

    if task["status"] == "failed":
        raise ValidationError(
            message=f"Job failed: {task.get('error', 'Unknown error')}",
            code="job_failed",
        )

    result = task.get("result") or {}
    file_id = result.get("file_id")
    if not file_id:
        raise NotFoundError(
            message="Export file not found",
            code="export_file_missing",
        )

    export = database.export_files.get_export_file(tenant_id, file_id)
    if not export:
        raise NotFoundError(
            message="Export file not found",
            code="export_file_missing",
        )

    database.export_files.mark_downloaded(tenant_id, file_id)

    backend = storage.get_backend()

    if export["storage_type"] == "spaces":
        url = backend.get_download_url(
            export["storage_path"],
            export["filename"],
            expires_in=3600,
        )
        return {
            "storage_type": "spaces",
            "url": url,
            "filename": export["filename"],
        }
    else:
        file_path = backend.get_file_path(export["storage_path"])
        if not file_path:
            raise NotFoundError(
                message="Export file not found on disk",
                code="export_file_missing",
            )
        return {
            "storage_type": "local",
            "path": file_path,
            "filename": export["filename"],
            "content_type": export.get("content_type", XLSX_CONTENT_TYPE),
        }


def create_upload_task(
    requesting_user: RequestingUser,
    file_data: bytes,
    password: str = "",
) -> dict | None:
    """Save uploaded spreadsheet to storage and create a background task.

    Validates the file is a readable XLSX with the expected columns, then
    saves it to the storage backend and enqueues a worker task.

    If the file is encrypted and a password is provided, decryption is
    attempted for validation. The original (encrypted) file is stored as-is,
    and the password is passed to the worker via the job payload.

    Authorization: Requires admin or super_admin role.

    Args:
        requesting_user: The user making the request
        file_data: Raw bytes of the uploaded XLSX file
        password: Optional password for encrypted files

    Returns:
        Dict with task id and created_at

    Raises:
        ValidationError: If file is invalid or has wrong columns
    """
    require_admin(requesting_user)
    track_activity(requesting_user["tenant_id"], requesting_user["id"])

    tenant_id = requesting_user["tenant_id"]

    # Validate file structure before saving
    _validate_xlsx(file_data, password)

    # Save to storage (encrypted file stored as-is)
    storage_key = f"uploads/{tenant_id}/bulk-update-{uuid4().hex[:12]}.xlsx"
    backend = storage.get_backend()
    backend.save(storage_key, BytesIO(file_data), XLSX_CONTENT_TYPE)

    storage_type = settings.STORAGE_BACKEND.lower()
    if storage_type != "spaces" or not settings.SPACES_BUCKET:
        storage_type = "local"

    payload: dict[str, Any] = {
        "storage_key": storage_key,
        "storage_type": storage_type,
    }
    if password:
        payload["password"] = password

    result = database.bg_tasks.create_task(
        tenant_id=tenant_id,
        job_type="bulk_update_users",
        created_by=requesting_user["id"],
        payload=payload,
    )

    if result:
        log_event(
            tenant_id=tenant_id,
            actor_user_id=requesting_user["id"],
            artifact_type="bg_task",
            artifact_id=str(result["id"]),
            event_type="bulk_update_task_created",
            metadata={"job_type": "bulk_update_users"},
        )

    return result


def _validate_xlsx(file_data: bytes, password: str = "") -> None:
    """Validate that file_data is a readable XLSX with expected columns.

    If the file cannot be opened directly and a password is provided,
    attempts decryption before validation. Raises a clear error if the
    file is encrypted but no password was given.
    """
    from openpyxl import load_workbook

    data_to_validate = file_data
    try:
        wb = load_workbook(filename=BytesIO(data_to_validate), read_only=True, data_only=True)
    except Exception as open_err:
        # OLE2 magic bytes indicate an encrypted OOXML file (Excel saves
        # encrypted XLSX as OLE2 containers, not ZIP). Plain XLSX starts with PK.
        is_encrypted = file_data[:4] == b"\xd0\xcf\x11\xe0"
        if not is_encrypted:
            raise ValidationError(
                message=f"Invalid XLSX file: {open_err}",
                code="invalid_file",
            )
        if not password:
            raise ValidationError(
                message="File is encrypted. Please provide the file password.",
                code="encrypted_file",
            )
        try:
            from utils.xlsx_encryption import decrypt_xlsx_data

            data_to_validate = decrypt_xlsx_data(file_data, password)
            wb = load_workbook(filename=BytesIO(data_to_validate), read_only=True, data_only=True)
        except ValidationError:
            raise
        except Exception as e:
            raise ValidationError(
                message=f"Failed to decrypt file. Check the password and try again. ({e})",
                code="decrypt_failed",
            )

    ws = wb.active
    if ws is None:
        wb.close()
        raise ValidationError(
            message="Spreadsheet has no active sheet",
            code="invalid_file",
        )

    rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
    wb.close()

    if not rows:
        raise ValidationError(
            message="Spreadsheet is empty",
            code="empty_file",
        )

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    if header != EXPECTED_COLUMNS:
        raise ValidationError(
            message=f"Invalid columns. Expected: {EXPECTED_COLUMNS}",
            code="invalid_columns",
        )
