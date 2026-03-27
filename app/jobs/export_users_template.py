"""Bulk user template export job handler."""

import logging
from datetime import UTC, datetime, timedelta
from io import BytesIO
from typing import Any

import database
import settings
from jobs.registry import register_handler
from utils import storage

logger = logging.getLogger(__name__)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TEMPLATE_COLUMNS = [
    "user_id",
    "email",
    "domain",
    "first_name",
    "last_name",
    "new_secondary_email",
    "new_first_name",
    "new_last_name",
]

# Columns 1-5 are locked reference data; 6-8 are editable
_LOCKED_COLS = 5
_TOTAL_COLS = 8


@register_handler("export_users_template")
def handle_export_users_template(task: dict) -> dict[str, Any]:
    """Generate an XLSX template of all active users for bulk update.

    Args:
        task: The task dict with id, tenant_id, job_type, payload, created_by

    Returns:
        Dict with output message, file_id, records_processed, filename, file_size
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Protection

    tenant_id = str(task["tenant_id"])
    created_by = str(task["created_by"])
    task_id = str(task["id"])

    logger.info("Starting user template export for tenant %s", tenant_id)

    tenant = database.tenants.get_tenant_by_id(tenant_id)
    if not tenant:
        raise ValueError(f"Tenant {tenant_id} does not exist")

    subdomain = tenant.get("subdomain", "users")

    users = database.users.list_all_users_for_export(tenant_id)
    logger.info("Found %d users to export", len(users))

    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    base_font = Font(size=16)
    header_font = Font(bold=True, size=16)
    locked_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    unlocked_prot = Protection(locked=False)

    # Header row
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        if col_idx <= _LOCKED_COLS:
            cell.fill = locked_fill

    # User data rows
    for row_idx, user in enumerate(users, start=2):
        email = user.get("email", "")
        domain = email.split("@")[1] if "@" in email else ""

        row_data = [
            str(user["id"]),
            email,
            domain,
            user.get("first_name", ""),
            user.get("last_name", ""),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = base_font
            cell.fill = locked_fill

        # Editable columns (6-8): set font, unlock
        for col_idx in range(_LOCKED_COLS + 1, _TOTAL_COLS + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = base_font
            cell.protection = unlocked_prot

    # Auto-filter on header row
    last_row = len(users) + 1
    ws.auto_filter.ref = f"A1:H{last_row}"

    # Sheet protection: lock reference cells, allow formatting and filtering
    ws.protection.sheet = True
    ws.protection.autoFilter = False
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.enable()

    # Column widths
    ws.column_dimensions["A"].width = 40  # user_id (UUID)
    ws.column_dimensions["B"].width = 35  # email
    ws.column_dimensions["C"].width = 25  # domain
    ws.column_dimensions["D"].width = 20  # first_name
    ws.column_dimensions["E"].width = 20  # last_name
    ws.column_dimensions["F"].width = 35  # new_secondary_email
    ws.column_dimensions["G"].width = 20  # new_first_name
    ws.column_dimensions["H"].width = 20  # new_last_name

    # Save to BytesIO
    file_buffer = BytesIO()
    wb.save(file_buffer)
    file_buffer.seek(0)
    file_size = file_buffer.getbuffer().nbytes

    # Generate filename and storage key
    timestamp = datetime.now(UTC).strftime("%Y-%m-%d_%H%M%S")
    filename = f"{subdomain}_users_{timestamp}.xlsx"
    storage_key = f"exports/{tenant_id}/{filename}"

    # Save to storage backend
    backend = storage.get_backend()
    backend.save(storage_key, file_buffer, XLSX_CONTENT_TYPE)

    storage_type = settings.STORAGE_BACKEND.lower()
    if storage_type != "spaces" or not settings.SPACES_BUCKET:
        storage_type = "local"

    expires_at = datetime.now(UTC) + timedelta(hours=settings.EXPORT_FILE_EXPIRY_HOURS)

    export_file = database.export_files.create_export_file(
        tenant_id=tenant_id,
        bg_task_id=task_id,
        filename=filename,
        storage_type=storage_type,
        storage_path=storage_key,
        file_size=file_size,
        expires_at=expires_at,
        created_by=created_by,
        content_type=XLSX_CONTENT_TYPE,
    )

    logger.info(
        "Created user template export: %s",
        export_file["id"] if export_file else "None",
    )

    size_kb = file_size // 1024
    output_msg = f"Generated user template with {len(users):,} users ({filename}, {size_kb:,} KB)"

    return {
        "output": output_msg,
        "file_id": str(export_file["id"]) if export_file else None,
        "records_processed": len(users),
        "filename": filename,
        "file_size": file_size,
    }
