"""Tests for Event Export job handler (XLSX format).

Tests cover:
- XLSX export with encryption and password in result
- Date range filtering via payload
- Artifact name resolution (users, groups, SPs, IdPs, OAuth2 clients)
- Actor email resolution
- Row limit guard (1M rows)
- Filename format with date ranges
- Batch pagination
"""

from datetime import UTC, date, datetime, timedelta
from typing import Any
from unittest.mock import MagicMock, patch
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from utils.xlsx_encryption import decrypt_xlsx_data


def _prepare_event_metadata(
    custom_metadata: dict[str, Any] | None = None,
) -> tuple[dict[str, Any], str]:
    """Helper to prepare combined_metadata and metadata_hash for create_event()."""
    from utils.request_metadata import compute_metadata_hash

    combined_metadata: dict[str, Any] = {
        "device": None,
        "remote_address": None,
        "session_id_hash": None,
        "user_agent": None,
    }
    if custom_metadata:
        combined_metadata.update(custom_metadata)

    metadata_hash = compute_metadata_hash(combined_metadata)
    return combined_metadata, metadata_hash


# =============================================================================
# XLSX Export Handler Tests
# =============================================================================


def test_export_produces_encrypted_xlsx(test_tenant, test_admin_user):
    """Test export produces an encrypted XLSX with password in result."""
    import database
    from jobs.export_events import handle_export_events

    # Create test events
    for i in range(3):
        combined_metadata, metadata_hash = _prepare_event_metadata({"index": i})
        database.event_log.create_event(
            tenant_id=test_tenant["id"],
            tenant_id_value=test_tenant["id"],
            event_type="test_event",
            artifact_type="test",
            artifact_id=str(uuid4()),
            actor_user_id=test_admin_user["id"],
            combined_metadata=combined_metadata,
            metadata_hash=metadata_hash,
        )

    bg_task = database.bg_tasks.create_task(
        tenant_id=test_tenant["id"],
        job_type="export_events",
        created_by=test_admin_user["id"],
    )

    task = {
        "id": bg_task["id"],
        "tenant_id": test_tenant["id"],
        "created_by": test_admin_user["id"],
        "job_type": "export_events",
    }

    captured_data = None

    def capture_save(storage_key, file_obj, content_type):
        nonlocal captured_data
        file_obj.seek(0)
        captured_data = file_obj.read()
        return storage_key

    with patch("jobs.export_events.storage.get_backend") as mock_get_backend:
        mock_backend = MagicMock()
        mock_backend.save.side_effect = capture_save
        mock_get_backend.return_value = mock_backend

        result = handle_export_events(task)

    # Verify result structure
    assert "output" in result
    assert "file_id" in result
    assert "password" in result
    assert "records_processed" in result
    assert "filename" in result
    assert "file_size" in result

    assert result["records_processed"] == 3
    assert result["password"]  # Non-empty password
    assert result["filename"].endswith(".xlsx")

    # Verify the file is encrypted and can be decrypted
    assert captured_data is not None
    decrypted = decrypt_xlsx_data(captured_data, result["password"])
    from io import BytesIO

    wb = load_workbook(BytesIO(decrypted))
    ws = wb.active
    assert ws.title == "Audit Log"

    # Verify headers
    headers = [cell.value for cell in ws[1]]
    assert headers[0] == "Timestamp"
    assert headers[1] == "Event Type"
    assert headers[2] == "Description"
    assert headers[3] == "Actor Email"
    assert headers[11] == "Metadata"

    # Verify data rows
    assert ws.max_row == 4  # 1 header + 3 data rows


def test_export_with_no_events(test_tenant, test_admin_user):
    """Test export with zero events produces headers-only XLSX."""
    import database
    from jobs.export_events import handle_export_events

    bg_task = database.bg_tasks.create_task(
        tenant_id=test_tenant["id"],
        job_type="export_events",
        created_by=test_admin_user["id"],
    )

    task = {
        "id": bg_task["id"],
        "tenant_id": test_tenant["id"],
        "created_by": test_admin_user["id"],
        "job_type": "export_events",
    }

    with patch("jobs.export_events.storage.get_backend") as mock_get_backend:
        mock_backend = MagicMock()
        mock_backend.save.return_value = f"exports/{test_tenant['id']}/test.xlsx"
        mock_get_backend.return_value = mock_backend

        result = handle_export_events(task)

    assert result["records_processed"] == 0
    assert "0 events" in result["output"]
    assert result["password"]
    assert result["file_id"] is not None


def test_export_with_date_range_payload():
    """Test that date range from payload is passed to list_events."""
    from jobs.export_events import handle_export_events

    tenant_id = str(uuid4())
    user_id = str(uuid4())
    task_id = str(uuid4())
    file_id = str(uuid4())

    task = {
        "id": task_id,
        "tenant_id": tenant_id,
        "created_by": user_id,
        "job_type": "export_events",
        "payload": {
            "start_date": "2026-01-01",
            "end_date": "2026-03-31",
        },
    }

    with (
        patch("jobs.export_events.database") as mock_db,
        patch("jobs.export_events.storage.get_backend") as mock_get_backend,
        patch("jobs.export_events.encrypt_workbook") as mock_encrypt,
    ):
        mock_db.tenants.get_tenant_by_id.return_value = {"id": tenant_id}
        mock_db.event_log.list_events.return_value = []
        mock_db.export_files.create_export_file.return_value = {"id": file_id}

        mock_encrypt.return_value = MagicMock(
            data=MagicMock(), password="test-pass", file_size=1024
        )

        mock_backend = MagicMock()
        mock_get_backend.return_value = mock_backend

        result = handle_export_events(task)

    # Verify list_events was called with date params
    call_kwargs = mock_db.event_log.list_events.call_args.kwargs
    assert call_kwargs["start_date"] == date(2026, 1, 1)
    assert call_kwargs["end_date"] == date(2026, 3, 31)

    # Verify filename contains date range
    assert "2026-01-01" in result["filename"]
    assert "2026-03-31" in result["filename"]


def test_export_pagination():
    """Test export processes multiple batches correctly."""
    from jobs.export_events import handle_export_events

    tenant_id = str(uuid4())
    user_id = str(uuid4())
    task_id = str(uuid4())
    file_id = str(uuid4())

    def make_event(i):
        return {
            "id": str(uuid4()),
            "tenant_id": tenant_id,
            "event_type": "test_event",
            "artifact_type": "test",
            "artifact_id": str(uuid4()),
            "actor_user_id": user_id,
            "created_at": datetime.now(UTC),
            "metadata": {},
        }

    batch1 = [make_event(i) for i in range(1000)]
    batch2 = [make_event(i) for i in range(500)]

    task = {
        "id": task_id,
        "tenant_id": tenant_id,
        "created_by": user_id,
        "job_type": "export_events",
    }

    with (
        patch("jobs.export_events.database") as mock_db,
        patch("jobs.export_events.storage.get_backend") as mock_get_backend,
        patch("jobs.export_events.encrypt_workbook") as mock_encrypt,
        patch("jobs.export_events._resolve_artifact_names", return_value={}),
        patch(
            "jobs.export_events._resolve_actor_emails",
            return_value={user_id: "test@example.com"},
        ),
    ):
        mock_db.tenants.get_tenant_by_id.return_value = {"id": tenant_id}
        mock_db.event_log.list_events.side_effect = [batch1, batch2, []]
        mock_db.export_files.create_export_file.return_value = {"id": file_id}

        mock_encrypt.return_value = MagicMock(
            data=MagicMock(), password="test-pass", file_size=1024
        )
        mock_backend = MagicMock()
        mock_get_backend.return_value = mock_backend

        result = handle_export_events(task)

    assert result["records_processed"] == 1500
    assert mock_db.event_log.list_events.call_count == 3  # 2 batches + 1 empty


def test_export_row_limit_guard():
    """Test export fails when exceeding 1M rows."""
    from jobs.export_events import handle_export_events

    tenant_id = str(uuid4())
    user_id = str(uuid4())

    # Create a batch of 1001 events to trigger on first batch
    # (we mock returning > MAX_EXPORT_ROWS total)
    big_batch = [
        {
            "id": str(uuid4()),
            "tenant_id": tenant_id,
            "event_type": "test",
            "artifact_type": "test",
            "artifact_id": str(uuid4()),
            "actor_user_id": user_id,
            "created_at": datetime.now(UTC),
            "metadata": {},
        }
        for _ in range(1000)
    ]

    task = {
        "id": str(uuid4()),
        "tenant_id": tenant_id,
        "created_by": user_id,
        "job_type": "export_events",
    }

    with (
        patch("jobs.export_events.database") as mock_db,
        patch("jobs.export_events.storage.get_backend"),
        patch("jobs.export_events.MAX_EXPORT_ROWS", 1500),
        patch("jobs.export_events._resolve_artifact_names", return_value={}),
        patch("jobs.export_events._resolve_actor_emails", return_value={}),
    ):
        mock_db.tenants.get_tenant_by_id.return_value = {"id": tenant_id}
        # Return 1000 events twice (total 2000 > 1500 limit)
        mock_db.event_log.list_events.side_effect = [big_batch, big_batch]

        with pytest.raises(ValueError, match="exceeds"):
            handle_export_events(task)


def test_export_nonexistent_tenant():
    """Test export raises when tenant does not exist."""
    from jobs.export_events import handle_export_events

    task = {
        "id": str(uuid4()),
        "tenant_id": str(uuid4()),
        "created_by": str(uuid4()),
        "job_type": "export_events",
    }

    with patch("jobs.export_events.storage.get_backend") as mock_get_backend:
        mock_backend = MagicMock()
        mock_get_backend.return_value = mock_backend

        with pytest.raises(ValueError, match="does not exist"):
            handle_export_events(task)

        assert not mock_backend.save.called


def test_export_artifact_name_resolution():
    """Test that artifact names are resolved for non-user types."""
    from jobs.export_events import _resolve_artifact_names

    tenant_id = str(uuid4())
    group_id = str(uuid4())
    sp_id = str(uuid4())

    events = [
        {"artifact_type": "group", "artifact_id": group_id},
        {"artifact_type": "service_provider", "artifact_id": sp_id},
        {"artifact_type": "user", "artifact_id": str(uuid4())},
    ]

    with patch("jobs.export_events.database") as mock_db:
        mock_db.groups.get_group_by_id.return_value = {"name": "Engineering"}
        mock_db.service_providers.get_service_provider.return_value = {"name": "Slack"}

        names = _resolve_artifact_names(tenant_id, events)

    assert names[("group", group_id)] == "Engineering"
    assert names[("service_provider", sp_id)] == "Slack"
    # Users are handled inline, not in this dict
    assert not any(k[0] == "user" for k in names)


def test_export_actor_email_resolution():
    """Test that actor emails are resolved from primary emails."""
    from jobs.export_events import _resolve_actor_emails
    from services.event_log import SYSTEM_ACTOR_ID

    tenant_id = str(uuid4())
    user_id = str(uuid4())

    events = [
        {"actor_user_id": user_id},
        {"actor_user_id": SYSTEM_ACTOR_ID},
    ]

    with patch("jobs.export_events.database") as mock_db:
        mock_db.user_emails.get_primary_email.return_value = {"email": "alice@example.com"}

        emails = _resolve_actor_emails(tenant_id, events)

    assert emails[user_id] == "alice@example.com"
    assert emails[SYSTEM_ACTOR_ID] == "system"


def test_export_filename_all():
    """Test filename for full export (no date range)."""
    from jobs.export_events import _build_filename

    filename = _build_filename(None, None)
    assert filename.startswith("audit-log_all_")
    assert filename.endswith(".xlsx")


def test_export_filename_date_range():
    """Test filename includes date range."""
    from jobs.export_events import _build_filename

    filename = _build_filename(date(2026, 1, 1), date(2026, 3, 31))
    assert "2026-01-01" in filename
    assert "2026-03-31" in filename
    assert filename.endswith(".xlsx")


def test_export_filename_start_only():
    """Test filename with start date only."""
    from jobs.export_events import _build_filename

    filename = _build_filename(date(2026, 1, 1), None)
    assert "2026-01-01" in filename
    assert "present" in filename


def test_export_filename_end_only():
    """Test filename with end date only."""
    from jobs.export_events import _build_filename

    filename = _build_filename(None, date(2026, 3, 31))
    assert "2026-03-31" in filename
    assert "up-to" in filename


def test_export_storage_type_local(test_tenant, test_admin_user):
    """Test export uses local storage type when not using Spaces."""
    import database
    from jobs.export_events import handle_export_events

    bg_task = database.bg_tasks.create_task(
        tenant_id=test_tenant["id"],
        job_type="export_events",
        created_by=test_admin_user["id"],
    )

    task = {
        "id": bg_task["id"],
        "tenant_id": test_tenant["id"],
        "created_by": test_admin_user["id"],
        "job_type": "export_events",
    }

    with patch("jobs.export_events.storage.get_backend") as mock_get_backend:
        with patch("jobs.export_events.settings.STORAGE_BACKEND", "local"):
            mock_backend = MagicMock()
            mock_backend.save.return_value = f"exports/{test_tenant['id']}/test.xlsx"
            mock_get_backend.return_value = mock_backend

            result = handle_export_events(task)

            export_file = database.export_files.get_export_file(
                test_tenant["id"], result["file_id"]
            )
            assert export_file["storage_type"] == "local"


def test_export_expiry_calculation(test_tenant, test_admin_user):
    """Test that export file expiry is calculated correctly."""
    import database
    from jobs.export_events import handle_export_events

    bg_task = database.bg_tasks.create_task(
        tenant_id=test_tenant["id"],
        job_type="export_events",
        created_by=test_admin_user["id"],
    )

    task = {
        "id": bg_task["id"],
        "tenant_id": test_tenant["id"],
        "created_by": test_admin_user["id"],
        "job_type": "export_events",
    }

    with patch("jobs.export_events.storage.get_backend") as mock_get_backend:
        with patch("jobs.export_events.settings.EXPORT_FILE_EXPIRY_HOURS", 48):
            mock_backend = MagicMock()
            mock_backend.save.return_value = f"exports/{test_tenant['id']}/test.xlsx"
            mock_get_backend.return_value = mock_backend

            before_export = datetime.now(UTC)
            result = handle_export_events(task)
            after_export = datetime.now(UTC)

            export_file = database.export_files.get_export_file(
                test_tenant["id"], result["file_id"]
            )
            expires_at = export_file["expires_at"]

            expected_min = before_export + timedelta(hours=48)
            expected_max = after_export + timedelta(hours=48)
            assert expected_min <= expires_at <= expected_max


def test_export_content_type(test_tenant, test_admin_user):
    """Test that export file has XLSX content type."""
    import database
    from jobs.export_events import XLSX_CONTENT_TYPE, handle_export_events

    bg_task = database.bg_tasks.create_task(
        tenant_id=test_tenant["id"],
        job_type="export_events",
        created_by=test_admin_user["id"],
    )

    task = {
        "id": bg_task["id"],
        "tenant_id": test_tenant["id"],
        "created_by": test_admin_user["id"],
        "job_type": "export_events",
    }

    with patch("jobs.export_events.storage.get_backend") as mock_get_backend:
        mock_backend = MagicMock()
        mock_backend.save.return_value = f"exports/{test_tenant['id']}/test.xlsx"
        mock_get_backend.return_value = mock_backend

        result = handle_export_events(task)

        export_file = database.export_files.get_export_file(test_tenant["id"], result["file_id"])
        assert export_file["content_type"] == XLSX_CONTENT_TYPE


def test_export_filename_has_timestamp():
    """Test filename includes a YYYYMMDD-HHMMSS timestamp."""
    from jobs.export_events import _build_filename

    filename = _build_filename(None, None)
    # audit-log_all_YYYYMMDD-HHMMSS_xxxxxxxx.xlsx
    parts = filename.replace("audit-log_all_", "").replace(".xlsx", "").split("_")
    assert len(parts) == 2  # timestamp + hex suffix
    assert len(parts[0]) == 15  # YYYYMMDD-HHMMSS
