"""Bulk user attribute update job handler.

Reads a previously uploaded XLSX file from storage, processes each row
to add secondary emails and/or update names, and returns a summary.
"""

import logging
from io import BytesIO
from typing import Any

import database
from jobs.registry import register_handler
from services.event_log import log_event
from utils import storage
from utils.request_context import system_context

logger = logging.getLogger(__name__)

EXPECTED_COLUMNS = [
    "user_id",
    "email",
    "domain",
    "first_name",
    "last_name",
    "new_secondary_email",
    "new_first_name",
    "new_last_name",
]

# Column indices (0-based)
_COL_USER_ID = 0
_COL_NEW_EMAIL = 5
_COL_NEW_FIRST = 6
_COL_NEW_LAST = 7
_NUM_COLS = 8


@register_handler("bulk_update_users")
def handle_bulk_update_users(task: dict) -> dict[str, Any]:
    """Process an uploaded spreadsheet to apply bulk user attribute updates.

    For each row:
    - If new_secondary_email is non-empty: add as verified secondary email
      (skip if address already exists in tenant)
    - If new_first_name or new_last_name is non-empty: update those fields
    - Rows with all blank "new" columns are skipped silently

    Args:
        task: The task dict with id, tenant_id, job_type, payload, created_by

    Returns:
        Dict with summary counts and per-row errors
    """
    from openpyxl import load_workbook

    tenant_id = str(task["tenant_id"])
    actor_id = str(task["created_by"])
    payload = task.get("payload") or {}
    storage_key = payload["storage_key"]

    logger.info("Starting bulk user update for tenant %s", tenant_id)

    # Read file from storage
    backend = storage.get_backend()
    file_path = backend.get_file_path(storage_key)
    if not file_path:
        raise ValueError(f"Upload file not found: {storage_key}")

    with open(file_path, "rb") as f:
        file_data = f.read()

    wb = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Skip header
    data_rows = rows[1:]
    logger.info("Processing %d rows", len(data_rows))

    emails_added = 0
    names_updated = 0
    rows_skipped = 0
    row_errors: list[dict[str, Any]] = []

    with system_context():
        for row_num, row in enumerate(data_rows, start=2):
            row_list = list(row) + [None] * max(0, _NUM_COLS - len(row))

            user_id = str(row_list[_COL_USER_ID]).strip() if row_list[_COL_USER_ID] else ""
            new_email = (
                str(row_list[_COL_NEW_EMAIL]).strip().lower()
                if row_list[_COL_NEW_EMAIL]
                else ""
            )
            new_first = (
                str(row_list[_COL_NEW_FIRST]).strip() if row_list[_COL_NEW_FIRST] else ""
            )
            new_last = (
                str(row_list[_COL_NEW_LAST]).strip() if row_list[_COL_NEW_LAST] else ""
            )

            if not new_email and not new_first and not new_last:
                rows_skipped += 1
                continue

            if not user_id:
                row_errors.append({"row": row_num, "error": "Missing user_id"})
                continue

            user = database.users.get_user_by_id(tenant_id, user_id)
            if not user:
                row_errors.append(
                    {"row": row_num, "user_id": user_id, "error": "User not found"}
                )
                continue

            if new_email:
                try:
                    _add_secondary_email(tenant_id, actor_id, user_id, new_email)
                    emails_added += 1
                except Exception as e:
                    row_errors.append(
                        {
                            "row": row_num,
                            "user_id": user_id,
                            "error": f"Email: {e}",
                        }
                    )

            if new_first or new_last:
                try:
                    if _update_name(tenant_id, actor_id, user_id, user, new_first, new_last):
                        names_updated += 1
                except Exception as e:
                    row_errors.append(
                        {
                            "row": row_num,
                            "user_id": user_id,
                            "error": f"Name: {e}",
                        }
                    )

        log_event(
            tenant_id=tenant_id,
            actor_user_id=actor_id,
            artifact_type="tenant",
            artifact_id=tenant_id,
            event_type="bulk_update_completed",
            metadata={
                "emails_added": emails_added,
                "names_updated": names_updated,
                "rows_skipped": rows_skipped,
                "errors": len(row_errors),
                "total_rows": len(data_rows),
            },
        )

    # Clean up the uploaded file
    try:
        backend.delete(storage_key)
    except Exception:
        logger.warning("Failed to clean up upload file: %s", storage_key)

    output_parts = []
    if emails_added:
        output_parts.append(f"{emails_added:,} emails added")
    if names_updated:
        output_parts.append(f"{names_updated:,} names updated")
    if rows_skipped:
        output_parts.append(f"{rows_skipped:,} rows skipped")
    if row_errors:
        output_parts.append(f"{len(row_errors):,} errors")
    output_msg = f"Processed {len(data_rows):,} rows: " + ", ".join(output_parts or ["no changes"])

    return {
        "output": output_msg,
        "emails_added": emails_added,
        "names_updated": names_updated,
        "rows_skipped": rows_skipped,
        "row_errors": row_errors,
        "total_rows": len(data_rows),
    }


def _add_secondary_email(
    tenant_id: str,
    actor_id: str,
    user_id: str,
    email: str,
) -> None:
    """Add a verified secondary email, skipping if already in use."""
    if database.user_emails.email_exists(tenant_id, email):
        raise ValueError(f"Email {email} already in use")

    result = database.user_emails.add_verified_email(
        tenant_id=tenant_id,
        user_id=user_id,
        email=email,
        tenant_id_value=tenant_id,
        is_primary=False,
    )

    if not result:
        raise ValueError("Failed to add email")

    log_event(
        tenant_id=tenant_id,
        actor_user_id=actor_id,
        artifact_type="user",
        artifact_id=user_id,
        event_type="email_added",
        metadata={
            "email": email,
            "email_id": str(result["id"]),
            "is_admin_action": True,
            "auto_verified": True,
            "bulk_update": True,
        },
    )


def _update_name(
    tenant_id: str,
    actor_id: str,
    user_id: str,
    current_user: dict,
    new_first: str,
    new_last: str,
) -> bool:
    """Update user name fields, using current values for blank fields.

    Returns True if any field actually changed, False if no-op.
    """
    changes: dict[str, dict[str, str]] = {}
    if new_first and new_first != current_user["first_name"]:
        changes["first_name"] = {
            "old": current_user["first_name"],
            "new": new_first,
        }
    if new_last and new_last != current_user["last_name"]:
        changes["last_name"] = {
            "old": current_user["last_name"],
            "new": new_last,
        }

    if not changes:
        return False

    first_name = new_first if new_first else current_user["first_name"]
    last_name = new_last if new_last else current_user["last_name"]

    database.users.update_user_profile(
        tenant_id=tenant_id,
        user_id=user_id,
        first_name=first_name,
        last_name=last_name,
    )

    log_event(
        tenant_id=tenant_id,
        actor_user_id=actor_id,
        artifact_type="user",
        artifact_id=user_id,
        event_type="user_updated",
        metadata={"changes": changes, "bulk_update": True},
    )

    return True
