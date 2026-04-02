"""Tests for User Audit Export job handler (multi-sheet XLSX).

Tests cover:
- 3-sheet workbook creation with correct headers
- Data population for each sheet
- Status/creation_method/auth_method mapping
- Secondary emails aggregation
- App access group attribution
- Empty data edge cases
- Encryption and password in result
"""

from datetime import UTC, datetime
from io import BytesIO
from typing import Any
from unittest.mock import MagicMock, patch
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from utils.xlsx_encryption import decrypt_xlsx_data


def _make_user(
    *,
    user_id: str | None = None,
    first_name: str = "Jane",
    last_name: str = "Doe",
    primary_email: str = "jane@example.com",
    role: str = "member",
    is_inactivated: bool = False,
    is_anonymized: bool = False,
    has_password: bool = True,
    mfa_enabled: bool = False,
    saml_idp_id: str | None = None,
    saml_idp_name: str | None = None,
    last_login: Any = None,
    last_activity_at: Any = None,
    password_changed_at: Any = None,
    created_at: Any = None,
) -> dict:
    return {
        "id": user_id or str(uuid4()),
        "first_name": first_name,
        "last_name": last_name,
        "primary_email": primary_email,
        "role": role,
        "is_inactivated": is_inactivated,
        "is_anonymized": is_anonymized,
        "has_password": has_password,
        "mfa_enabled": mfa_enabled,
        "saml_idp_id": saml_idp_id,
        "saml_idp_name": saml_idp_name,
        "last_login": last_login,
        "last_activity_at": last_activity_at,
        "password_changed_at": password_changed_at,
        "created_at": created_at or datetime.now(UTC),
    }


def _run_handler(users, **kwargs):
    """Run the export handler with mocked database and storage, return result and workbook."""
    import database

    tenant_id = str(uuid4())
    user_id = str(uuid4())
    task_id = str(uuid4())

    secondary_emails = kwargs.get("secondary_emails", [])
    creation_methods = kwargs.get("creation_methods", [])
    login_ips = kwargs.get("login_ips", [])
    app_counts = kwargs.get("app_counts", [])
    ata_count = kwargs.get("ata_count", 0)
    memberships = kwargs.get("memberships", [])
    access_rows = kwargs.get("access_rows", [])
    assertions = kwargs.get("assertions", [])

    # Create tenant
    tenant = database.tenants.get_tenant_by_id(tenant_id)
    if not tenant:
        # Mock the tenant lookup
        pass

    task = {
        "id": task_id,
        "tenant_id": tenant_id,
        "created_by": user_id,
        "job_type": "export_users",
    }

    captured_data = None

    def capture_save(storage_key, file_obj, content_type):
        nonlocal captured_data
        file_obj.seek(0)
        captured_data = file_obj.read()
        return storage_key

    with (
        patch("jobs.export_users.database") as mock_db,
        patch("jobs.export_users.storage.get_backend") as mock_get_backend,
        patch("jobs.export_users.settings") as mock_settings,
    ):
        mock_db.tenants.get_tenant_by_id.return_value = {"id": tenant_id}
        mock_db.users.get_all_users_for_audit.return_value = users
        mock_db.users.get_all_secondary_emails.return_value = secondary_emails
        mock_db.users.get_creation_methods.return_value = creation_methods
        mock_db.users.get_last_login_ips.return_value = login_ips
        mock_db.users.get_app_counts.return_value = app_counts
        mock_db.users.get_available_to_all_sp_count.return_value = ata_count
        mock_db.users.get_all_group_memberships.return_value = memberships
        mock_db.users.get_all_user_sp_access.return_value = access_rows
        mock_db.users.get_last_sso_assertions.return_value = assertions
        mock_db.export_files.create_export_file.return_value = {"id": str(uuid4())}

        mock_backend = MagicMock()
        mock_backend.save.side_effect = capture_save
        mock_get_backend.return_value = mock_backend
        mock_settings.STORAGE_BACKEND = "local"
        mock_settings.SPACES_BUCKET = ""
        mock_settings.EXPORT_FILE_EXPIRY_HOURS = 24

        from jobs.export_users import handle_export_users

        result = handle_export_users(task)

    assert captured_data is not None
    decrypted = decrypt_xlsx_data(captured_data, result["password"])
    wb = load_workbook(BytesIO(decrypted))

    return result, wb


# =============================================================================
# Basic workbook structure
# =============================================================================


def test_export_produces_three_sheets():
    """Export produces a 3-sheet encrypted XLSX."""
    user = _make_user()
    result, wb = _run_handler([user])

    assert len(wb.sheetnames) == 3
    assert wb.sheetnames == ["Users", "Group Memberships", "App Access"]

    assert "output" in result
    assert "password" in result
    assert "file_id" in result
    assert "filename" in result
    assert result["filename"].endswith(".xlsx")
    assert result["records_processed"] == 1


def test_export_with_no_users():
    """Export with zero users produces headers-only sheets."""
    result, wb = _run_handler([])

    assert result["records_processed"] == 0
    ws_users = wb["Users"]
    assert ws_users.max_row == 1  # Header only


# =============================================================================
# Sheet 1: Users
# =============================================================================


def test_users_sheet_headers():
    """Users sheet has correct headers."""
    _, wb = _run_handler([_make_user()])
    ws = wb["Users"]
    headers = [cell.value for cell in ws[1]]

    assert headers == [
        "User ID",
        "First Name",
        "Last Name",
        "Primary Email",
        "Domain",
        "Secondary Emails",
        "Role",
        "Status",
        "Created",
        "Creation Method",
        "Auth Method",
        "Last Login",
        "Last Login IP",
        "Last Activity",
        "Password Changed",
        "Two-Step Verification",
        "App Count",
    ]


def test_users_sheet_status_active():
    """Active user shows status 'active'."""
    user = _make_user()
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=8).value == "active"


def test_users_sheet_status_inactive():
    """Inactivated user shows status 'inactive'."""
    user = _make_user(is_inactivated=True)
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=8).value == "inactive"


def test_users_sheet_status_anonymized():
    """Anonymized user shows status 'anonymized'."""
    user = _make_user(is_anonymized=True)
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=8).value == "anonymized"


def test_users_sheet_creation_method_invited():
    """User with user_created event shows 'invited'."""
    uid = str(uuid4())
    user = _make_user(user_id=uid)
    _, wb = _run_handler(
        [user],
        creation_methods=[{"user_id": uid, "event_type": "user_created"}],
    )
    ws = wb["Users"]
    assert ws.cell(row=2, column=10).value == "invited"


def test_users_sheet_creation_method_jit():
    """User with user_created_jit event shows 'jit'."""
    uid = str(uuid4())
    user = _make_user(user_id=uid)
    _, wb = _run_handler(
        [user],
        creation_methods=[{"user_id": uid, "event_type": "user_created_jit"}],
    )
    ws = wb["Users"]
    assert ws.cell(row=2, column=10).value == "jit"


def test_users_sheet_creation_method_cli():
    """User with no creation event shows 'cli'."""
    user = _make_user()
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=10).value == "cli"


def test_users_sheet_auth_method_password():
    """User with password shows 'password'."""
    user = _make_user(has_password=True)
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=11).value == "password"


def test_users_sheet_auth_method_idp():
    """User with IdP shows IdP name."""
    user = _make_user(saml_idp_name="Okta", has_password=False)
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=11).value == "Okta"


def test_users_sheet_auth_method_none():
    """User with no password and no IdP shows 'none'."""
    user = _make_user(has_password=False)
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=11).value == "none"


def test_users_sheet_mfa_enabled():
    """MFA enabled shows 'Yes'."""
    user = _make_user(mfa_enabled=True)
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=16).value == "Yes"


def test_users_sheet_mfa_disabled():
    """MFA disabled shows 'No'."""
    user = _make_user(mfa_enabled=False)
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=16).value == "No"


def test_users_sheet_secondary_emails():
    """Secondary emails appear comma-separated."""
    uid = str(uuid4())
    user = _make_user(user_id=uid)
    _, wb = _run_handler(
        [user],
        secondary_emails=[
            {"user_id": uid, "email": "alt1@example.com"},
            {"user_id": uid, "email": "alt2@example.com"},
        ],
    )
    ws = wb["Users"]
    assert ws.cell(row=2, column=6).value == "alt1@example.com, alt2@example.com"


def test_users_sheet_domain_extracted():
    """Domain is extracted from primary email."""
    user = _make_user(primary_email="admin@acme.com")
    _, wb = _run_handler([user])
    ws = wb["Users"]
    assert ws.cell(row=2, column=5).value == "acme.com"


def test_users_sheet_app_count():
    """App count includes group-based and available_to_all SPs."""
    uid = str(uuid4())
    user = _make_user(user_id=uid)
    _, wb = _run_handler(
        [user],
        app_counts=[{"user_id": uid, "app_count": 3}],
        ata_count=2,
    )
    ws = wb["Users"]
    assert ws.cell(row=2, column=17).value == 5  # 3 group-based + 2 ata


def test_users_sheet_last_login_ip():
    """Last login IP extracted from event log metadata."""
    uid = str(uuid4())
    user = _make_user(user_id=uid)
    _, wb = _run_handler(
        [user],
        login_ips=[{"user_id": uid, "metadata": {"remote_address": "10.0.0.1"}}],
    )
    ws = wb["Users"]
    assert ws.cell(row=2, column=13).value == "10.0.0.1"


# =============================================================================
# Sheet 2: Group Memberships
# =============================================================================


def test_memberships_sheet_headers():
    """Group Memberships sheet has correct headers."""
    _, wb = _run_handler([_make_user()])
    ws = wb["Group Memberships"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["User ID", "Email", "Group Name", "Group Type", "Member Since"]


def test_memberships_sheet_data():
    """Membership rows appear in the sheet."""
    uid = str(uuid4())
    now = datetime.now(UTC)
    _, wb = _run_handler(
        [_make_user(user_id=uid)],
        memberships=[
            {
                "user_id": uid,
                "email": "jane@example.com",
                "group_name": "Engineering",
                "group_type": "weftid",
                "membership_since": now,
            },
        ],
    )
    ws = wb["Group Memberships"]
    assert ws.max_row == 2  # Header + 1 data row
    assert ws.cell(row=2, column=3).value == "Engineering"
    assert ws.cell(row=2, column=4).value == "weftid"


# =============================================================================
# Sheet 3: App Access
# =============================================================================


def test_access_sheet_headers():
    """App Access sheet has correct headers."""
    _, wb = _run_handler([_make_user()])
    ws = wb["App Access"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["User ID", "Email", "App Name", "Last Auth", "Access Via"]


def test_access_sheet_group_attribution():
    """Access via groups shows comma-separated group names."""
    uid = str(uuid4())
    sp_id = str(uuid4())
    _, wb = _run_handler(
        [_make_user(user_id=uid)],
        access_rows=[
            {
                "user_id": uid,
                "email": "jane@example.com",
                "sp_id": sp_id,
                "app_name": "Jira",
                "available_to_all": False,
                "granting_group_name": "Engineering",
            },
            {
                "user_id": uid,
                "email": "jane@example.com",
                "sp_id": sp_id,
                "app_name": "Jira",
                "available_to_all": False,
                "granting_group_name": "DevOps",
            },
        ],
    )
    ws = wb["App Access"]
    assert ws.max_row == 2  # Header + 1 aggregated row
    assert ws.cell(row=2, column=3).value == "Jira"
    access_via = ws.cell(row=2, column=5).value
    assert "DevOps" in access_via
    assert "Engineering" in access_via


def test_access_sheet_available_to_all():
    """Available-to-all SP shows 'All users'."""
    uid = str(uuid4())
    sp_id = str(uuid4())
    _, wb = _run_handler(
        [_make_user(user_id=uid)],
        access_rows=[
            {
                "user_id": uid,
                "email": "jane@example.com",
                "sp_id": sp_id,
                "app_name": "Slack",
                "available_to_all": True,
                "granting_group_name": None,
            },
        ],
    )
    ws = wb["App Access"]
    assert ws.cell(row=2, column=5).value == "All users"


def test_access_sheet_last_auth():
    """Last auth timestamp from SSO assertion events."""
    uid = str(uuid4())
    sp_id = str(uuid4())
    now = datetime.now(UTC)
    _, wb = _run_handler(
        [_make_user(user_id=uid)],
        access_rows=[
            {
                "user_id": uid,
                "email": "jane@example.com",
                "sp_id": sp_id,
                "app_name": "Jira",
                "available_to_all": False,
                "granting_group_name": "Engineering",
            },
        ],
        assertions=[
            {"user_id": uid, "sp_id": sp_id, "last_auth_at": now},
        ],
    )
    ws = wb["App Access"]
    assert ws.cell(row=2, column=4).value != ""


# =============================================================================
# Nonexistent tenant
# =============================================================================


def test_export_nonexistent_tenant():
    """Export raises ValueError for nonexistent tenant."""
    task = {
        "id": str(uuid4()),
        "tenant_id": str(uuid4()),
        "created_by": str(uuid4()),
        "job_type": "export_users",
    }

    with patch("jobs.export_users.database") as mock_db:
        mock_db.tenants.get_tenant_by_id.return_value = None

        from jobs.export_users import handle_export_users

        with pytest.raises(ValueError, match="does not exist"):
            handle_export_users(task)


def test_fmt_dt_string_passthrough():
    """_fmt_dt passes through non-datetime, non-None values as strings."""
    from jobs.export_users import _fmt_dt

    assert _fmt_dt("2026-01-01") == "2026-01-01"
    assert _fmt_dt(42) == "42"
